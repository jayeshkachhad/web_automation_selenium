import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains, Keys
from openpyxl import load_workbook
from openpyxl.styles import Font
# from openpyxl.styles import colors
# from openpyxl.cell import Cell

wb = load_workbook('contacts.xlsx')
ws = wb.active

driver = webdriver.Chrome()
actions = ActionChains(driver)
url = 'https://web.whatsapp.com/'

driver.get(url)
driver.maximize_window()
time.sleep(40)

maxRow = len(ws['A'])

sent = 0


try:
    for i in range(1, maxRow):
        pick_contact = ws.cell(row=i, column=1).value
        actions.send_keys(Keys.ESCAPE)
        actions.send_keys(Keys.ESCAPE)
        driver.find_element(By.CSS_SELECTOR, ".selectable-text.copyable-text.x15bjb6t.x1n2onr6").send_keys(pick_contact)
        time.sleep(3)

        try:
            try:
                driver.find_element(By.CSS_SELECTOR, "._ak8l").click()
                time.sleep(2)
            except Exception as e:
                driver.find_element(By.CSS_SELECTOR, "span[data-icon='x-alt']").click()
                print(f"{pick_contact} It is not Whatsapp Contact")
                ws.cell(row=i, column=1).font = Font(color='FFFF0000')
                ws.cell(row=i, column=2).value = str(e)
                time.sleep(2)
                actions.send_keys(Keys.ESCAPE)
                continue
            driver.find_element(By.XPATH, "//div[@class='_ajyw']").click()
            driver.find_element(By.XPATH, "(//p[@class='selectable-text copyable-text x15bjb6t x1n2onr6'])[2]").click()
            driver.find_element(By.XPATH, "(//p[@class='selectable-text copyable-text x15bjb6t x1n2onr6'])[2]").send_keys("💸💸 EARN MORE MONEY 💸💸 \n💰 1. No need to leave your present job/bussiness \n 💰 2. No Target Pressure \n 💰 3. No Time Boundation \n 💰 4. Yes, Highly Reputed Brand \n 💰 5. Yes, Consumer Satisfaction \n 💰 6. Yes, Work Life Balance \n 💰 7. Yes, Impressive income \n WhatsApp at \n http://wa.me/917990439256 \n Contact :- 7990439256")
            # driver.find_element(By.CSS_SELECTOR, "div[title='Type a message'] p[class='selectable-text copyable-text x15bjb6t x1n2onr6']").send_keys("💸💸 EARN MORE MONEY 💸💸 \n💰 1. No need to leave your present job/bussiness \n 💰 2. No Target Pressure \n 💰 3. No Time Boundation \n 💰 4. Yes, Highly Reputed Brand \n 💰 5. Yes, Consumer Satisfaction \n 💰 6. Yes, Work Life Balance \n 💰 7. Yes, Impressive income \n WhatsApp at \n http://wa.me/917990439256 \n Contact :- 7990439256")
            # actions.key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()
            # time.sleep(1)
            time.sleep(2.5)
            driver.find_element(By.CSS_SELECTOR, "span[data-icon='send']").click()
            # actions.send_keys(Keys.ENTER)
            # driver.find_element(By.CSS_SELECTOR, "span[data-icon='send']").click()
            time.sleep(2)
            ws.cell(row=i, column=1).font = Font(color="00ff00")
            print(f"{pick_contact} is sent successfully = {sent}")
            ws.cell(row=i, column=1).font = Font(color="00ff00")
            sent = sent + 1
            ws.cell(row=i, column=2).value = "Sent " + str(sent)
            wb.save("contacts.xlsx")

        except Exception as e:
            actions.key_down(Keys.ESCAPE)
            time.sleep(1)
            actions.key_down(Keys.ESCAPE)
            # time.sleep(1)
            actions.key_down(Keys.ESCAPE)
            # time.sleep(1)
            continue

    print(f"{sent} contacts sent")
except Exception as e:
    ws.cell(row=2, column=1).value = f"Program failed with {sent} contacts"
    print(e)
    # os.system("shutdown /s /t 1")
