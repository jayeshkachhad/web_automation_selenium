import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains, Keys
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

wb = load_workbook('contacts.xlsx')
ws = wb.active
maxRow = len(ws['A'])

driver = webdriver.Chrome()
actions = ActionChains(driver)
url = 'https://web.whatsapp.com/'

driver.get(url)
driver.maximize_window()
time.sleep(40)