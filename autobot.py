import time
# import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains, Keys
from openpyxl import load_workbook
# from openpyxl.styles import Font

wb = load_workbook('Ahmedabadi.xlsx')
ws = wb.active

driver = webdriver.Chrome()
actions = ActionChains(driver)
url = 'https://web.telegram.org/k'

driver.get(url)
driver.maximize_window()
time.sleep(60)

maxRow = len(ws['A'])
print(maxRow)

sent = 0

ws.cell(row=1, column=1)

try:
    for i in range(1, maxRow):
        pick_contact = ws.cell(row=i, column=2).value
        # print(pick_contact)
        if i % 100 == 0:
            time.sleep(240)

        try:
            driver.find_element(By.CSS_SELECTOR, "button[class='btn-icon rp btn-menu-toggle sidebar-tools-button is-visible'] div[class='c-ripple']").click()
            time.sleep(1)
            driver.find_element(By.XPATH,"(//div[@class='btn-menu-item rp-overflow'])[3]").click()
            time.sleep(1)
            driver.find_element(By.XPATH, "(//input[contains(@placeholder,'')])[2]").send_keys(pick_contact)
            time.sleep(2)
            try:
                driver.find_element(By.CSS_SELECTOR, "body > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > div:nth-child(1) > div:nth-child(2) > ul:nth-child(1) > a:nth-child(1) > div:nth-child(1)").click()
                time.sleep(1)
            except:
                time.sleep(2)
                actions.send_keys(Keys.ESCAPE)
                # driver.find_element(By.CSS_SELECTOR, ".btn-icon.sidebar-back-button.is-visible").click()
                ws.cell(row=i, column=2).value = "0"
                continue
            # driver.find_element(By.XPATH, "//div[@class='input-message-input is-empty scrollable scrollable-y no-scrollbar']").send_keys("http://lipowers.online")
            actions.key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()
            time.sleep(1)
            # actions.send_keys(Keys.ENTER)
            driver.find_element(By.CSS_SELECTOR, "button[class='btn-icon rp btn-circle btn-send animated-button-icon send']").click()
            # driver.find_element(By.CSS_SELECTOR, "button[class='btn-icon rp btn-circle btn-send animated-button-icon send'] div[class='c-ripple']").click()
            time.sleep(4)
            ws.cell(row=i, column=2).value = "1"
            ws.cell(row=i, column=3).value = sent
            sent = sent + 1
            wb.save("Ahmedabad.xlsx")
            print(f"sent to {pick_contact}")
        except:
            ws.cell(row=i, column=2).value = "0"
            wb.save("Ahmedabad.xlsx")
            time.sleep(1)
            continue
except:
    print("Stopped")
    # ws.cell(row=i, column=2).value = "0"