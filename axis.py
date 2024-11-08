import time

from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.styles import Font
from plyer import notification
# from pushbullet import Pushbullet
#
#
# API_KEY = "o.QVhTjKwCECLRDhVEUPR8PJNhrIAgxipO"
# pb = Pushbullet(API_KEY)

wb = load_workbook('urnlist.xlsx')
ws = wb.active

driver = webdriver.Chrome()
act = ActionChains(driver)

url = "https://nseitexams.com/OnlinePayment/paymentRequest.jsp?5617db59-15b2-40f2-be44-fb2c716d2375"

driver.get(url)
driver.maximize_window()
time.sleep(5)

start_ = 909240001
strL = "LICI0"

for i in range(10000):
    start_ = start_ + 1
    urnlink = strL + str(start_)
    # print(urnlink)
    driver.find_element(By.XPATH, "//input[@id='txtTransId']").clear()
    driver.find_element(By.XPATH, "//input[@id='txtTransId']").send_keys(urnlink)
    driver.find_element(By.XPATH, "//input[@name='btnSearch']").click()
    time.sleep(1)
    # driver.switchTo().alert().dismiss()
    driver.switch_to.alert.dismiss()
    time.sleep(1)
    try:
        nameis = driver.find_element(By.CSS_SELECTOR, "tbody tr td:nth-child(3)").text
        print(urnlink + " = " + nameis)
        if nameis.__contains__("RATHOD"):
            notification.notify("Alert", str(nameis))
            print("FOUND =================")
            # push = pb.push_note('Alert', 'Its rathod')
        else:
            pass
        ws.cell(row=i, column=1).value = urnlink
        ws.cell(row=i, column=2).value = str(nameis)
        wb.save("urnlist.xlsx")
    except:
        # print("Not Found")
        pass
    time.sleep(1)